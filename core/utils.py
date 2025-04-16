import jwt
from datetime import datetime, timedelta
from django.conf import settings
from typing import Dict, Any, Optional, Union
from django.utils import timezone
import io
import csv
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill
from django.http import HttpResponse

def generate_jwt_token(user_id: int, expires_delta: Optional[timedelta] = None) -> str:
    """Gera um token JWT para o usuário."""
    payload = {
        'user_id': user_id,
        'exp': datetime.utcnow() + (expires_delta or timedelta(hours=1)),
        'iat': datetime.utcnow(),
    }
    return jwt.encode(payload, settings.JWT_SECRET_KEY, algorithm='HS256')

def generate_tokens_for_user(user_id: int) -> Dict[str, str]:
    """Gera tokens de acesso e atualização para o usuário."""
    access_token = generate_jwt_token(
        user_id=user_id,
        expires_delta=timedelta(seconds=settings.JWT_ACCESS_TOKEN_LIFETIME)
    )
    
    refresh_token = generate_jwt_token(
        user_id=user_id,
        expires_delta=timedelta(seconds=settings.JWT_REFRESH_TOKEN_LIFETIME)
    )
    
    return {
        'access_token': access_token,
        'refresh_token': refresh_token,
        'token_type': 'bearer'
    }

def verify_jwt_token(token: str) -> Optional[Dict[str, Any]]:
    """Verifica a validade de um token JWT."""
    try:
        payload = jwt.decode(token, settings.JWT_SECRET_KEY, algorithms=['HS256'])
        return payload
    except jwt.ExpiredSignatureError:
        return None
    except jwt.InvalidTokenError:
        return None

def get_client_ip(request) -> str:
    """Obtém o endereço IP do cliente."""
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip

def generate_excel_file(data: list, headers: list, sheet_name: str = 'Sheet1') -> io.BytesIO:
    """Gera um arquivo Excel com os dados fornecidos."""
    output = io.BytesIO()
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    
    # Define estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Adiciona cabeçalhos
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        cell = sheet[f"{col_letter}1"]
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Adiciona dados
    for row_num, row_data in enumerate(data, 2):
        for col_num, value in enumerate(row_data, 1):
            sheet.cell(row=row_num, column=col_num).value = value
    
    # Ajusta largura das colunas
    for col_num in range(1, len(headers) + 1):
        sheet.column_dimensions[get_column_letter(col_num)].width = 15
    
    workbook.save(output)
    output.seek(0)
    
    return output

def generate_csv_file(data: list, headers: list) -> io.StringIO:
    """Gera um arquivo CSV com os dados fornecidos."""
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Adiciona cabeçalhos
    writer.writerow(headers)
    
    # Adiciona dados
    writer.writerows(data)
    
    output.seek(0)
    return output

def create_file_response(file_data: Union[io.BytesIO, io.StringIO], filename: str, file_type: str) -> HttpResponse:
    """Cria uma resposta HTTP com um arquivo para download."""
    if file_type == 'excel':
        response = HttpResponse(
            file_data.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
    elif file_type == 'csv':
        response = HttpResponse(file_data.getvalue(), content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
    else:
        raise ValueError(f"Tipo de arquivo não suportado: {file_type}")
    
    return response

def format_datetime(dt: datetime, include_time: bool = True) -> str:
    """Formata um objeto datetime para exibição."""
    if include_time:
        return dt.strftime('%d/%m/%Y %H:%M')
    return dt.strftime('%d/%m/%Y')

def is_business_hour(dt: datetime) -> bool:
    """Verifica se o datetime está dentro do horário comercial (8h às 18h)."""
    return 8 <= dt.hour < 18 and dt.weekday() < 5  # Segunda a sexta

def add_business_days(date_val: datetime, num_days: int) -> datetime:
    """Adiciona um número de dias úteis a uma data."""
    current_date = date_val
    added_days = 0
    
    while added_days < num_days:
        current_date += timedelta(days=1)
        if current_date.weekday() < 5:  # Segunda a sexta
            added_days += 1
            
    return current_date 