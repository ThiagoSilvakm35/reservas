from celery import shared_task
from django.utils import timezone
from django.contrib.auth import get_user_model
from django.db.models import Count, Avg, Sum, Q, F
from datetime import timedelta, datetime
import logging
import io
import openpyxl
from openpyxl.chart import BarChart, Reference, PieChart
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill

from reservas.models import Booking, Provider
from avaliacoes.models import Review
from notificacoes.tasks import criar_e_enviar_notificacao
from .models import Report

User = get_user_model()
logger = logging.getLogger(__name__)

@shared_task
def enviar_relatorio_semanal():
    """Gera e envia um relatório semanal para administradores."""
    agora = timezone.now()
    inicio_semana = agora - timedelta(days=7)
    
    # Busca usuários administradores
    admins = User.objects.filter(user_type='admin', is_active=True)
    
    for admin in admins:
        try:
            # Verifica a frequência de relatórios preferida
            try:
                if admin.preferences.report_frequency != 'weekly':
                    continue
            except:
                pass
            
            # Gera o relatório
            titulo = f"Relatório Semanal - {inicio_semana.strftime('%d/%m/%Y')} a {agora.strftime('%d/%m/%Y')}"
            relatorio = gerar_relatorio_semanal(admin.id, inicio_semana, agora, titulo)
            
            # Envia notificação com o relatório
            criar_e_enviar_notificacao.delay(
                recipient_id=admin.id,
                notification_type='report',
                object_id=str(relatorio.id),
                entity_type='Report',
                context={
                    'report': {
                        'title': titulo,
                        'start_date': inicio_semana.strftime('%d/%m/%Y'),
                        'end_date': agora.strftime('%d/%m/%Y')
                    }
                }
            )
            
        except Exception as e:
            logger.error(f"Erro ao gerar relatório para {admin.email}: {str(e)}")
    
    return f"Processo de geração de relatórios semanais concluído para {admins.count()} administradores."

@shared_task
def gerar_relatorio_semanal(admin_id, inicio, fim, titulo):
    """Gera um relatório semanal."""
    try:
        admin = User.objects.get(id=admin_id)
        
        # Estatísticas de reservas
        reservas = Booking.objects.filter(
            created_at__gte=inicio,
            created_at__lte=fim
        )
        
        total_reservas = reservas.count()
        reservas_por_status = dict(reservas.values('status').annotate(total=Count('id')).values_list('status', 'total'))
        
        # Reservas por prestador
        reservas_por_prestador = list(reservas.values('provider__user__first_name', 'provider__user__last_name', 'provider__service_name')
                               .annotate(total=Count('id'))
                               .order_by('-total')[:10])
        
        # Reservas por dia
        reservas_por_dia = list(reservas.values('start_datetime__date')
                         .annotate(data=F('start_datetime__date'), total=Count('id'))
                         .values('data', 'total')
                         .order_by('data'))
        
        # Avaliações
        avaliacoes = Review.objects.filter(
            created_at__gte=inicio,
            created_at__lte=fim
        )
        
        total_avaliacoes = avaliacoes.count()
        media_avaliacoes = avaliacoes.aggregate(Avg('rating'))['rating__avg'] or 0
        
        # Gera um relatório em Excel
        report_file = _gerar_excel_relatorio(
            titulo=titulo,
            inicio=inicio,
            fim=fim,
            total_reservas=total_reservas,
            reservas_por_status=reservas_por_status,
            reservas_por_prestador=reservas_por_prestador,
            reservas_por_dia=reservas_por_dia,
            total_avaliacoes=total_avaliacoes,
            media_avaliacoes=media_avaliacoes
        )
        
        # Cria o objeto Report
        report = Report.objects.create(
            title=titulo,
            type='weekly',
            format='excel',
            start_date=inicio.date(),
            end_date=fim.date(),
            created_by=admin,
            data={
                'total_reservas': total_reservas,
                'reservas_por_status': reservas_por_status,
                'total_avaliacoes': total_avaliacoes,
                'media_avaliacoes': media_avaliacoes
            }
        )
        
        # Salva o arquivo
        report.file.save(f"relatorio_semanal_{inicio.strftime('%Y%m%d')}.xlsx", report_file)
        
        return report
    
    except User.DoesNotExist:
        logger.error(f"Administrador {admin_id} não encontrado.")
        raise
    except Exception as e:
        logger.error(f"Erro ao gerar relatório semanal: {str(e)}")
        raise

def _gerar_excel_relatorio(titulo, inicio, fim, total_reservas, reservas_por_status,
                          reservas_por_prestador, reservas_por_dia, 
                          total_avaliacoes, media_avaliacoes):
    """Gera um arquivo Excel com o relatório."""
    output = io.BytesIO()
    workbook = openpyxl.Workbook()
    
    # Remove a planilha padrão e cria novas
    workbook.remove(workbook.active)
    resumo = workbook.create_sheet("Resumo")
    detalhes = workbook.create_sheet("Detalhes")
    
    # Estilo dos títulos
    titulo_font = Font(bold=True, size=14, color="FFFFFF")
    titulo_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    
    # Estilo dos cabeçalhos
    cabecalho_font = Font(bold=True, color="FFFFFF")
    cabecalho_fill = PatternFill(start_color="8DB4E2", end_color="8DB4E2", fill_type="solid")
    
    # Planilha de Resumo
    resumo['A1'] = titulo
    resumo.merge_cells('A1:F1')
    resumo['A1'].font = titulo_font
    resumo['A1'].fill = titulo_fill
    resumo['A1'].alignment = Alignment(horizontal="center")
    
    resumo['A3'] = "Período:"
    resumo['B3'] = f"{inicio.strftime('%d/%m/%Y')} a {fim.strftime('%d/%m/%Y')}"
    resumo['A3'].font = Font(bold=True)
    
    resumo['A5'] = "Total de Reservas:"
    resumo['B5'] = total_reservas
    resumo['A5'].font = Font(bold=True)
    
    resumo['A6'] = "Reservas Pendentes:"
    resumo['B6'] = reservas_por_status.get('pending', 0)
    
    resumo['A7'] = "Reservas Confirmadas:"
    resumo['B7'] = reservas_por_status.get('confirmed', 0)
    
    resumo['A8'] = "Reservas Canceladas:"
    resumo['B8'] = reservas_por_status.get('canceled', 0)
    
    resumo['A9'] = "Reservas Concluídas:"
    resumo['B9'] = reservas_por_status.get('completed', 0)
    
    resumo['A11'] = "Total de Avaliações:"
    resumo['B11'] = total_avaliacoes
    resumo['A11'].font = Font(bold=True)
    
    resumo['A12'] = "Média de Avaliações:"
    resumo['B12'] = f"{media_avaliacoes:.1f} / 5.0"
    
    # Gráfico de pizza para status de reservas
    pie = PieChart()
    labels = Reference(resumo, min_col=1, min_row=6, max_row=9)
    data = Reference(resumo, min_col=2, min_row=6, max_row=9)
    pie.add_data(data)
    pie.set_categories(labels)
    pie.title = "Reservas por Status"
    
    resumo.add_chart(pie, "D5")
    
    # Planilha de Detalhes
    detalhes['A1'] = "Reservas por Prestador"
    detalhes.merge_cells('A1:D1')
    detalhes['A1'].font = titulo_font
    detalhes['A1'].fill = titulo_fill
    detalhes['A1'].alignment = Alignment(horizontal="center")
    
    # Cabeçalhos
    detalhes['A3'] = "Prestador"
    detalhes['B3'] = "Serviço"
    detalhes['C3'] = "Total de Reservas"
    
    for col in ['A', 'B', 'C']:
        detalhes[f'{col}3'].font = cabecalho_font
        detalhes[f'{col}3'].fill = cabecalho_fill
    
    # Dados de reservas por prestador
    for i, item in enumerate(reservas_por_prestador, 4):
        detalhes[f'A{i}'] = f"{item['provider__user__first_name']} {item['provider__user__last_name']}"
        detalhes[f'B{i}'] = item['provider__service_name']
        detalhes[f'C{i}'] = item['total']
    
    # Seção de Reservas por Dia
    linha_inicio = i + 2
    detalhes[f'A{linha_inicio}'] = "Reservas por Dia"
    detalhes.merge_cells(f'A{linha_inicio}:C{linha_inicio}')
    detalhes[f'A{linha_inicio}'].font = titulo_font
    detalhes[f'A{linha_inicio}'].fill = titulo_fill
    detalhes[f'A{linha_inicio}'].alignment = Alignment(horizontal="center")
    
    # Cabeçalhos
    linha_cabecalho = linha_inicio + 2
    detalhes[f'A{linha_cabecalho}'] = "Data"
    detalhes[f'B{linha_cabecalho}'] = "Total de Reservas"
    
    for col in ['A', 'B']:
        detalhes[f'{col}{linha_cabecalho}'].font = cabecalho_font
        detalhes[f'{col}{linha_cabecalho}'].fill = cabecalho_fill
    
    # Dados de reservas por dia
    for j, item in enumerate(reservas_por_dia, linha_cabecalho + 1):
        detalhes[f'A{j}'] = item['data'].strftime('%d/%m/%Y')
        detalhes[f'B{j}'] = item['total']
    
    # Ajusta largura das colunas
    for worksheet in [resumo, detalhes]:
        for col in worksheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            worksheet.column_dimensions[column].width = max_length + 4
    
    workbook.save(output)
    output.seek(0)
    
    return output

@shared_task
def gerar_estatisticas_dashboard():
    """Gera estatísticas para o dashboard administrativo."""
    try:
        # Período para estatísticas
        agora = timezone.now()
        inicio_mes = agora.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        inicio_semana = agora - timedelta(days=agora.weekday())
        inicio_semana = inicio_semana.replace(hour=0, minute=0, second=0, microsecond=0)
        
        # Contagens gerais
        total_usuarios = User.objects.count()
        total_prestadores = Provider.objects.count()
        total_reservas = Booking.objects.count()
        
        # Reservas por status
        reservas_por_status = dict(Booking.objects.values('status').annotate(total=Count('id')).values_list('status', 'total'))
        
        # Reservas por dia nos últimos 30 dias
        ultimos_30_dias = agora - timedelta(days=30)
        reservas_por_dia = list(Booking.objects.filter(created_at__gte=ultimos_30_dias)
                         .values('start_datetime__date')
                         .annotate(data=F('start_datetime__date'), total=Count('id'))
                         .values('data', 'total')
                         .order_by('data'))
        
        # Reservas por hora
        reservas_por_hora = list(Booking.objects.filter(created_at__gte=ultimos_30_dias)
                          .values('start_datetime__hour')
                          .annotate(hora=F('start_datetime__hour'), total=Count('id'))
                          .values('hora', 'total')
                          .order_by('hora'))
        
        # Top prestadores
        top_prestadores = list(Booking.objects.filter(created_at__gte=inicio_mes)
                        .values('provider__id', 'provider__user__first_name', 'provider__user__last_name', 'provider__service_name')
                        .annotate(total=Count('id'))
                        .order_by('-total')[:5])
        
        # Prestadores melhor avaliados
        top_avaliados = list(Review.objects.filter(created_at__gte=inicio_mes)
                      .values('booking__provider__id', 'booking__provider__user__first_name', 'booking__provider__user__last_name')
                      .annotate(media=Avg('rating'), total=Count('id'))
                      .filter(total__gte=5)  # Pelo menos 5 avaliações
                      .order_by('-media')[:5])
        
        # Distribuição de avaliações
        distribuicao_avaliacoes = dict(Review.objects.values('rating').annotate(total=Count('id')).values_list('rating', 'total'))
        
        return {
            'total_usuarios': total_usuarios,
            'total_prestadores': total_prestadores,
            'total_reservas': total_reservas,
            'reservas_por_status': reservas_por_status,
            'reservas_por_dia': reservas_por_dia,
            'reservas_por_hora': reservas_por_hora,
            'top_prestadores': top_prestadores,
            'top_avaliados': top_avaliados,
            'distribuicao_avaliacoes': distribuicao_avaliacoes,
            'gerado_em': agora.isoformat(),
            'periodo': {
                'inicio_mes': inicio_mes.isoformat(),
                'inicio_semana': inicio_semana.isoformat(),
                'fim': agora.isoformat()
            }
        }
    
    except Exception as e:
        logger.error(f"Erro ao gerar estatísticas do dashboard: {str(e)}")
        return {
            'erro': str(e)
        } 